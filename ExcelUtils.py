import openpyxl

# 'S.No', 'Proposal Details[Proposal No,File No,Proposal Name]', 'Location[State,District,Tehsil]',
# 'Important Dates [Date of Submission for TOR,Proposal Accepted by MS on,Date of TOR Granted ]',TOR, 'Category', 'Company/Proponent',
#  'Type of project', '* Attached Files', '490070820151sardaenergy45-12.pdf', 'Cover letter',
#  'EC', 'EIA File', 'Additional Information', 'EcReport'


def get_row_count(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_row


def get_column_count(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_column


def read_data(file, sheet_name, row_num, col_num):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.cell(row=row_num, column=col_num).value


def write_data(file, sheet_name, row_num, col_num, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(sheet_name)
    sheet.cell(row=row_num, column=col_num).value = data
    workbook.save(file)


def search_text_combination(search_text):
    length = len(search_text)
    combination_list = []
    for i, word in enumerate(search_text):
        combination_list.append(word)
        for nextword in search_text[i + 1:]:
            combination_list.append(combination_list[-1] + " " + nextword)
    return combination_list[:length]


def dataCorrection(data):
    newdict = {}
    for d in data:
        if data[d].find(':') != -1:
            if data[d].find('://') > -1:
                newdict[d] = data[d]
            else:
                temp = data[d].split('\n')
                for p in range(len(temp)):
                    newdict[temp[p].split(':')[0]] = temp[p].split(':')[1]
        else:
            newdict[d] = data[d]
    return newdict


def removeDuplicate(l):
    seen = set()
    new_l = []
    for d in l:
        t = tuple(d.items())
        if t not in seen:
            seen.add(t)
            new_l.append(d)
    return new_l

# tmp_word = "Sun Pharma Company Ltd"
# search_output = (search_text_combination(tmp_word.split()))
# print(search_output[::-1])
