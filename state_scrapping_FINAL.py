import os
import time
import pymssql
import datetime
import platform
import logging
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook  # reading/loading workbooks
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

# Here we are defining the browser properties
chrome_path = Service(ChromeDriverManager().install())
option = webdriver.ChromeOptions()
option.add_argument("--start-maximized")
option.add_argument("--disable-geolocation")
option.add_argument("--ignore-certificate-errors")
option.add_argument("--disable-popup-blocking")
option.add_argument("--disable-translate")
driver = webdriver.Chrome(service=chrome_path, options=option)


def collectData(driver, table_count, y):
    outer_rows = driver.find_elements(
        By.XPATH, "//*[@id='ctl00_ContentPlaceHolder1_GridView1']/tbody/tr")
    outer_cols = driver.find_elements(
        By.XPATH, "//*[@id='ctl00_ContentPlaceHolder1_GridView1']/tbody/tr/th")
    # table_count = len(driver.find_elements(
    #     By.XPATH, '//table[@id="ctl00_ContentPlaceHolder1_GridView1"]/tbody/tr['+str(len(outer_rows))+']/td'))
    row_count = 0
    if table_count != 1:
        row_count = len(outer_rows)+1
    else:
        row_count = len(outer_rows)

    print(row_count-2, " DATA FOUND FOR YEAR -- ", y)
    log.info("{} DATA FOUND FOR YEAR -- {}". format(row_count-2, y))
    for r in range(2, row_count):
        for c in range(1, len(outer_cols)-1):
            if r == row_count-1 and table_count == 1:
                print("Excluding the pagination Row")
                break  # Excluding the pagination row
            raw_data = []
            # Proposal Details
            raw_data.insert(0, str(driver.find_elements(By.XPATH,
                                                        '//*[@id="ctl00_ContentPlaceHolder1_GridView1"]/tbody/tr[' + str(
                                                            r) + ']/td[2]')[0].text).replace('\n', ','))
            # Location
            raw_data.insert(1, str(driver.find_elements(By.XPATH,
                                                        '//*[@id="ctl00_ContentPlaceHolder1_GridView1"]/tbody/tr[' + str(
                                                            r) + ']/td[3]')[0].text).replace('\n', ','))
            # Important Dates
            raw_data.insert(2, str(driver.find_elements(By.XPATH,
                                                        '//*[@id="ctl00_ContentPlaceHolder1_GridView1"]/tbody/tr[' + str(
                                                            r) + ']/td[4]')[0].text).replace('\n', ','))
            # Category
            raw_data.insert(3, driver.find_elements(By.XPATH,
                                                    '//*[@id="ctl00_ContentPlaceHolder1_GridView1"]/tbody/tr[' + str(
                                                        r) + ']/td[5]')[0].text)
            # Company/Proponent
            raw_data.insert(4, driver.find_elements(By.XPATH,
                                                    '//*[@id="ctl00_ContentPlaceHolder1_GridView1"]/tbody/tr[' + str(
                                                        r) + ']/td[6]')[0].text)
            # Current status
            raw_data.insert(5, driver.find_elements(By.XPATH,
                                                    '//*[@id="ctl00_ContentPlaceHolder1_GridView1"]/tbody/tr[' + str(
                                                        r) + ']/td[7]')[0].text)
            # Attached Files
            imageLinks = driver.find_elements(By.XPATH,
                                              '//*[@id="ctl00_ContentPlaceHolder1_GridView1"]/tbody/tr['+str(r)+']/td[8]//a')
            file_key = []
            file_link = []
            for link in imageLinks:
                f = link.get_attribute("title")
                file_key.append(f)
                lst = link.get_attribute("href")
                file_link.append(lst)
                # for idx, item in enumerate(file_key):
                #     if "TOR" in item:
                #         file_key[idx] = "TOR"
                #     if "EC" in item:
                #         file_key[idx] = "EC"
                #     if "EC Report" in item:
                #         file_key[idx] = "EcReport"
            ack_link = ''
            for i in range(len(file_key)):
                ack_link = ack_link + \
                    file_key[i]+"~" + file_link[i]+","
            # Attached Files
            raw_data.insert(6, ack_link)
            # Acknowlegment
            # raw_data.insert(11, ' ')
            ws.append(raw_data)
            wb.save(path)
            break  # Breaking the column


def stateDataScrapping(ws):
    # Here we are browsing the below website
    driver.get("http://environmentclearance.nic.in/searchproposal.aspx")
    # After Landing into website click on radioButton
    driver.find_element(
        By.ID, "ctl00_ContentPlaceHolder1_RadioButtonList1_2").click()
    # collecting all state links
    state_links = driver.find_elements(By.XPATH,
                                       '//*[@id="ctl00_ContentPlaceHolder1_Label1"]/div/a')
    # lst = []
    # c = 1
    # for state in state_links:
    #     if c >= 28:
    #         lst.append(state.get_attribute("href"))  # Getting All state Link
    #     # lst.append(state.get_attribute("href"))  # Getting All state Link
    #     c = c+1

    lst = []
    # iterate and collect the hyperlink of the state
    for state in state_links:
        lst.append(state.get_attribute("href"))  # Getting All state Link

# start iteration on state hyperlink
    for l in lst:
        print("LINK -- ", l)
        log.info("SCRAPPING START FOR  = {}". format(l))
        # clicking on the state link
        driver.get(l)
        # switching into the ifram to collect sumittedEC count
        driver.switch_to.frame(driver.find_element(By.TAG_NAME, 'iframe'))
        sumittedEC = driver.find_elements(
            By.XPATH, '//body/form/div[3]/ul/li[2]/a[2]')
        if sumittedEC[0].text != '0':
            click_value = sumittedEC[0].get_attribute('href')
            # Clicking on the number of submitted EC Report
            driver.get(click_value)
            yearOption = []
            sele = Select(driver.find_element(
                By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_ddlyear"]'))
            opt = sele.options
            year = []
            # collecting the years that start from 2018
            for i in range(1, len(opt)):
                if opt[i].text >= "2018":
                    year.append(opt[i].text)
            # start iteration for each year
            for y in year:
                # print("YEAR--", y)
                s = Select(driver.find_element(
                    By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_ddlyear"]'))
                # setting the year
                s.select_by_value(y)
                # clicking the search button
                opt = (s.first_selected_option).text
                # Waiting for javascript loader
                driver.find_element(
                    By.ID, "ctl00_ContentPlaceHolder1_btn").click()
                WebDriverWait(driver, 300).until_not(
                    EC.visibility_of_element_located(("id", "lightover11")))

                no_data_found = len(driver.find_elements(
                    By.XPATH, '//table[@id="ctl00_ContentPlaceHolder1_GridView1"]/tbody/tr[2]/td'))
                # last_rows will contain the pagination
                last_rows = driver.find_elements(
                    By.XPATH, "//*[@id='ctl00_ContentPlaceHolder1_GridView1']/tbody/tr")
                # counting number of tables
                table_count = len(driver.find_elements(
                    By.XPATH, '//table[@id="ctl00_ContentPlaceHolder1_GridView1"]/tbody/tr['+str(len(last_rows))+']/td'))
                if no_data_found == 1:
                    log.info("NO DATA FOUND FOR  YEAR = {}". format(y))
                    continue
                else:
                    # page has no pagination collect the first page data
                    collectData(driver, table_count, y)
                    # if page has pagination then
                    if table_count == 1:
                        page_number = len(driver.find_elements(
                            By.XPATH, '//table[@id="ctl00_ContentPlaceHolder1_GridView1"]/tbody/tr['+str(len(last_rows))+']/td/table/tbody/tr/td'))

                        for i in range(2, page_number+1):
                            print(" LAST ROW -- ", len(last_rows), " I -- ", i)
                            # collecting the number of pages link
                            page_link = driver.find_elements(
                                By.XPATH, '//table[@id="ctl00_ContentPlaceHolder1_GridView1"]/tbody/tr['+str(len(last_rows))+']/td/table/tbody/tr/td['+str(i)+']/a')
                            # print("PAGE LINK -- ", page_link)
                            if len(page_link) != 0:
                                # clicking the page link
                                page_link[0].click()
                                # wait for javascript loader
                                WebDriverWait(driver, 300).until_not(
                                    EC.visibility_of_element_located(("id", "lightover11")))
                                last_rows = driver.find_elements(
                                    By.XPATH, "//*[@id='ctl00_ContentPlaceHolder1_GridView1']/tbody/tr")
                                print("Clicked On -- ", i)
                                # Start Scrapping
                                collectData(driver, table_count, y)
                            else:
                                log.info(
                                    "MISS DATA FOR  YEAR = {} and Page = {}". format(y, i))
                                print("MISS DATA FOR  YEAR = ",
                                      y, "and Page = ", i)


def dumpToDatabase(path, row_count):
    log.info("** -- CONNECTING TO DATABASE -- **")
    conn = pymssql.connect(
        host=r'101.53.148.84',
        port=r'8649',
        user=r'user_at',
        password=r'Ye4@83kv3',
        database='db_at'
    )
    cursor = conn.cursor(as_dict=True)
    log.info("** -- DATABASE CONNECTION ESTABLISHED SUCCESSFULLY -- **")

    # workbook object is created
    wb_obj = load_workbook(path)

    sheet_obj = wb_obj.active
    max_row = len(sheet_obj['A'])
    max_col = row_count
    l = []
    for r in range(2, max_row+1):
        col = []
        for c in range(1, max_col + 1):
            cell_obj = sheet_obj.cell(row=r, column=c)
            col.insert(c, cell_obj.value)
            # print(cell_obj.value, end = ",")
        l.append(tuple(col))
    # print(l)

    log.info("INSERTION OF {} Records is STARTED Please Wait !!!". format(max_row))
    start_time = time.time()
    q = "insert into parivesh_state(Proposal_Details, Location ,Important_Dates, Category, Company_or_Proponent,Current_Status, Attached_Files) values(%s,%s,%s,%s,%s,%s,%s)"
    cursor.executemany(q, l)
    conn.commit()
    log.info("DATA IS COMMITTED TO DATABASE. Time Taken to Store {} Record is {}".format(
        max_row, time.time()-start_time))
    conn.close()


# ====== RUN Start from Here
start_time = time.time()
x = datetime.datetime.now()
# Here we are defining the name of the file...
today = "stateData_"+str(x.day) + str(x.month) + str(x.year)+"_" + \
    str(x.strftime("%H"))+str(x.strftime("%M"))+str(x.strftime("%S"))
ROOT_DIR = os.path.dirname(os.path.abspath(__file__))
# To get the Operating System
plat_os = platform.uname()[0]  # To get the Operating System
folder_path = ''
log_path = ''
error_path = ''
# Below logic will create folder in project root directory based on the operating system where we are running the code
if plat_os == 'Windows':
    folder_path = ROOT_DIR+"\\state_output\\"
    log_path = ROOT_DIR+"\\state_log\\"
    error_path = ROOT_DIR+"\\error_log\\"
else:
    folder_path = ROOT_DIR+"/state_output/"
    log_path = ROOT_DIR+"/state_log/"
    error_path = ROOT_DIR+"/error_log/"

if not os.path.exists(folder_path):
    os.makedirs(folder_path)
if not os.path.exists(log_path):
    os.makedirs(log_path)
if not os.path.exists(error_path):
    os.makedirs(error_path)

print(log_path+today+'.log')
# LOG_FORMAT = '{lineno} ***  {asctime} ----- {message}'
# log.basicConfig(filename=log_path+today+'.log',
#                 style='{', format=LOG_FORMAT, level=log.DEBUG)

# Creating INFO LOG file
formatter = logging.Formatter('%(asctime)s %(levelname)s %(message)s')


def setup_logger(name, log_file, level=logging.INFO):
    handler = logging.FileHandler(log_file)
    handler.setFormatter(formatter)
    logger = logging.getLogger(name)
    logger.setLevel(level)
    logger.addHandler(handler)
    return logger


# Info  File logger
log = setup_logger('first_logger', log_path+today+'.log')
log.info('This is just info Log')

# Error file logger
err_log = setup_logger('second_logger', error_path+today+'.log')
err_log.error('This is an error Log')


path = folder_path+today+".xlsx"


workbook = Workbook()
workbook.save(path)
wb = load_workbook(filename=path)
ws = wb['Sheet']
# Here we are defining the header name of excel file and saving it.
header_list = ["Proposal Details", "Location", "Important Dates",
               "Category", "Company/Proponent", "Current Status", "Attached Files"]
ws.append(header_list)
wb.save(path)
# This function will start scrapping the website data and store it in excel file
stateDataScrapping(ws)
# this function will dump the excel data into database
dumpToDatabase(path, len(header_list))
log.info("TOTAL TIME ----- {}".format(time.time()-start_time))
