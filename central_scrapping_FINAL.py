import os
import platform
import pymssql
import logging
import datetime
import time
import threading as Thread
from pip import main
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
import ExcelUtils
import pandas as pd
import numpy as np
from openpyxl import Workbook  # reading/loading workbooks
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from concurrent.futures import ThreadPoolExecutor


mydata = []
dictionary_1 = {}


def collectData(driver, table_count, tmp_word, page_count, ws, subsidiary_name):
    # setting rflag to check company matching or not
    rflag = 0
    paging = 'no'
    # collecting number of rows
    rows = len(driver.find_elements(
        By.XPATH, '//table[@class="table Grid1"]/tbody/tr'))
    if page_count == 1:
        end_row = rows+1
    else:
        end_row = rows-1
        paging = 'yes'
        log.info(
            "Pagination for the Company ----- {}".format(tmp_word))
    # Collecting number of columns
    cols = len(driver.find_elements(
        By.XPATH, '//table[@class="table Grid1"]/tbody/tr/th'))
    for row in range(2, end_row):
        for col in range(1, cols + 1):
            raw_data = []
            pos = row
            # if there is no pagination then start reading from first row
            if table_count == 0:
                # checking if no data found
                if driver.find_elements(By.XPATH,
                                        '//*[@id="ctl00_ContentPlaceHolder1_grdevents"]/tbody/tr[' + str(row) + ']/td')[0].text == "No Records Found":
                    break
                else:
                    proposal_status = driver.find_elements(By.XPATH,
                                                           '//*[@id="ctl00_ContentPlaceHolder1_grdevents"]/tbody/tr[' + str(
                                                               row) + ']/td[6]')[0].text
                    # collecting name of project
                    result = driver.find_elements(By.XPATH,
                                                  '//*[@id="ctl00_ContentPlaceHolder1_grdevents"]/tbody/tr[' + str(
                                                      row) + ']/td[5]')[0].text
                    # collecting name of company
                    result1 = driver.find_elements(By.XPATH,
                                                   '//*[@id="ctl00_ContentPlaceHolder1_grdevents"]/tbody/tr[' + str(
                                                       row) + ']/td[4]')[0].text
            # if there is pagination then start reading from  row+1
            else:
                if row < rows-1:
                    if driver.find_elements(By.XPATH,
                                            '//*[@id="ctl00_ContentPlaceHolder1_grdevents"]/tbody/tr[' + str(row+1) + ']/td')[0].text == "No Records Found":
                        # print("ELSE NO DATA FOUND")
                        break
                    else:
                        proposal_status = driver.find_elements(By.XPATH,
                                                               '//*[@id="ctl00_ContentPlaceHolder1_grdevents"]/tbody/tr[' + str(
                                                                   row+1) + ']/td[6]')[0].text
                        result = driver.find_elements(By.XPATH,
                                                      '//*[@id="ctl00_ContentPlaceHolder1_grdevents"]/tbody/tr[' + str(
                                                          row+1) + ']/td[5]')[0].text
                        result1 = driver.find_elements(By.XPATH,
                                                       '//*[@id="ctl00_ContentPlaceHolder1_grdevents"]/tbody/tr[' + str(
                                                           row+1) + ']/td[4]')[0].text
                        pos = row+1
            # print("SERIAL NUMBER===> ", driver.find_elements(By.XPATH,
            #                                                  '//*[@id="ctl00_ContentPlaceHolder1_grdevents"]/tbody/tr[' + str(
            #                                                      pos) + ']/td[1]')[0].text, "  ROW===>", row)

            # Decorating the company name
            final_tmp = tmp_word.lower().replace("and", "&")
            final_tmp = final_tmp.lower().replace("limited", "ltd")
            final_result = result.lower().replace("and", "&")
            final_result = final_result.lower().replace("limited", "ltd")
            # checking if the name of company or project matching to the inputed company name
            if final_tmp in final_result.lower() or final_tmp in result1.lower():
                # if matches then set rflag value to 1
                rflag = 1
                # if there is no pagination then start reading from first row
                if table_count == 0:
                    # click_item will get the href value of magnifier
                    if rows > 2:
                        click_item = driver.find_elements(By.XPATH,
                                                          '//*[@id="ctl00_ContentPlaceHolder1_grdevents"]/tbody/tr[' + str(
                                                              row) + ']/td[7]/headerstyle/itemstyle/headerstyle/itemtemplate/a/img')
                    else:
                        click_item = driver.find_elements(By.XPATH,
                                                          '//*[@id="ctl00_ContentPlaceHolder1_grdevents"]/tbody/tr/td/headerstyle/itemstyle/headerstyle/itemtemplate/a/img')
                # if there is pagination then start reading from  row+1
                else:
                    if row < rows-1:
                        if rows > 2:
                            click_item = driver.find_elements(By.XPATH,
                                                              '//*[@id="ctl00_ContentPlaceHolder1_grdevents"]/tbody/tr[' + str(
                                                                  row+1) + ']/td[7]/headerstyle/itemstyle/headerstyle/itemtemplate/a/img')
                        else:
                            click_item = driver.find_elements(By.XPATH,
                                                              '//*[@id="ctl00_ContentPlaceHolder1_grdevents"]/tbody/tr/td/headerstyle/itemstyle/headerstyle/itemtemplate/a/img')
                # Proposal No
                raw_data.insert(0, driver.find_elements(By.XPATH,
                                                        '//*[@id="ctl00_ContentPlaceHolder1_grdevents"]/tbody/tr[' + str(pos) + ']/td[2]')[0].text)
                # MOEFCC File No-3
                raw_data.insert(1, driver.find_elements(By.XPATH,
                                                        '//*[@id="ctl00_ContentPlaceHolder1_grdevents"]/tbody/tr[' + str(pos) + ']/td[3]')[0].text)
                # Project Name-4
                raw_data.insert(2, driver.find_elements(By.XPATH,
                                                        '//*[@id="ctl00_ContentPlaceHolder1_grdevents"]/tbody/tr[' + str(pos) + ']/td[4]')[0].text)
                # Company-5
                raw_data.insert(3, driver.find_elements(By.XPATH,
                                                        '//*[@id="ctl00_ContentPlaceHolder1_grdevents"]/tbody/tr[' + str(pos) + ']/td[5]')[0].text)
                # Proposal Status-6
                raw_data.insert(4, driver.find_elements(By.XPATH,
                                                        '//*[@id="ctl00_ContentPlaceHolder1_grdevents"]/tbody/tr[' + str(pos) + ']/td[6]')[0].text)
                if proposal_status == 'DELETED':
                    log.info(
                        "DELETED Status  for Company ----- {}".format(tmp_word))
                    break
                else:
                    WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable(click_item[0])).click()

                ll = driver.current_url
                # Here we are checking that page containing Acknowledge page if it is then keep the url else get the second page tabler data
                if 'PrintPreviews_report.aspx' not in ll:

                    outer = len(driver.find_elements(
                        By.XPATH, "//*[@class='table Grid1']/tbody/tr[2]/td"))
                    outer_rows = len(driver.find_elements(
                        By.XPATH, "//*[@class='table Grid1']/tbody/tr"))
                    # checking no data found
                    if outer == 1:
                        log.info(
                            "NO DATA FOUND for Company ----- {}".format(tmp_word))
                        print("NO DATA FOUND-----------")
                        driver.back()
                        break
                    # here excluding the last four row that are irrelevant
                    if len(driver.find_elements(By.XPATH, "//*[@class='table Grid1']/tbody/tr/th")) >= 12:
                        outer_cols = len(driver.find_elements(
                            By.XPATH, "//*[@class='table Grid1']/tbody/tr/th")) - 4
                    else:
                        outer_cols = len(driver.find_elements(
                            By.XPATH, "//*[@class='table Grid1']/tbody/tr/th"))

                    for r in range(2, outer_rows + 1):
                        # Start data collection from second page table
                        # Location
                        raw_data.insert(5, str(driver.find_elements(By.XPATH,
                                                                    '//*[@class="table Grid1"]/tbody/tr[' + str(
                                                                        r) + ']/td[3]')[0].text).replace('\n', ','))
                        # Important Dates
                        raw_data.insert(6, str(driver.find_elements(By.XPATH,
                                                                    '//*[@class="table Grid1"]/tbody/tr[' + str(
                                                                        r) + ']/td[4]')[0].text).replace('\n', ','))
                        # Category
                        raw_data.insert(7, driver.find_elements(By.XPATH,
                                                                '//*[@class="table Grid1"]/tbody/tr[' + str(
                                                                    r) + ']/td[5]')[0].text)
                        # Company/Proponent
                        raw_data.insert(8, driver.find_elements(By.XPATH,
                                                                '//*[@class="table Grid1"]/tbody/tr[' + str(
                                                                    r) + ']/td[6]')[0].text)
                        # Type of Project
                        raw_data.insert(9, driver.find_elements(By.XPATH,
                                                                '//*[@class="table Grid1"]/tbody/tr[' + str(
                                                                    r) + ']/td[7]')[0].text)
                        # collecting all link from Attachment column
                        imageLinks = driver.find_elements(By.XPATH,
                                                          '//*[@class="table Grid1"]/tbody/tr/td[8]//a')

                        file_key = []
                        file_link = []
                        for link in imageLinks:
                            f = link.get_attribute("title")
                            file_key.append(f)
                            lst = link.get_attribute("href")
                            file_link.append(lst)
                            for idx, item in enumerate(file_key):
                                if "TOR" in item:
                                    file_key[idx] = "TOR"
                                if "EC" in item:
                                    file_key[idx] = "EC"
                                if "EC Report" in item:
                                    file_key[idx] = "EcReport"
                        ack_link = ''
                        for i in range(len(file_key)):
                            ack_link = ack_link + \
                                file_key[i]+"~" + file_link[i]+","
                        # Attached Files
                        raw_data.insert(10, ack_link)
                        # Acknowlegment
                        raw_data.insert(11, ' ')
                        raw_data.insert(12, paging)
                        raw_data.insert(13, tmp_word)
                        raw_data.insert(14, subsidiary_name)
                        ws.append(raw_data)
                        wb.save(path)
                else:
                    # Acknowlegment
                    raw_data.insert(5, '')
                    raw_data.insert(6, '')
                    raw_data.insert(7, '')
                    raw_data.insert(8, '')
                    raw_data.insert(9, '')
                    raw_data.insert(10, '')
                    raw_data.insert(11, ll)
                    raw_data.insert(12, paging)
                    raw_data.insert(13, tmp_word)
                    raw_data.insert(14, subsidiary_name)
                    ws.append(raw_data)
                    wb.save(path)
                driver.back()
            break
    # returning the rflag value to check company name found or not
    return rflag
    # break
    # print("RAW_DATA=======", raw_data)


def centralDataScrapping(ws):
    # Here we are defining the browser properties
    chrome_path = Service(ChromeDriverManager().install())
    option = webdriver.ChromeOptions()
    option.add_argument("--start-maximized")
    option.add_argument("--disable-geolocation")
    option.add_argument("--ignore-certificate-errors")
    option.add_argument("--disable-popup-blocking")
    option.add_argument("--disable-translate")
    # option.add_argument("--headless")
    driver = webdriver.Chrome(service=chrome_path, options=option)
    # driver.set_window_position(-10000, 0)
# Here we are browsing the below website
    driver.get("http://environmentclearance.nic.in/searchproposal.aspx")
# After Landing into website click on radioButton
    driver.find_element(
        By.ID, "ctl00_ContentPlaceHolder1_RadioButtonList1_1").click()
    try:
        gtotal = 0
        # Reading the input file
        companyName = pd.read_excel("inputData.xlsx")
        list_of_company = companyName['Company Name'].tolist()
        list_of_subsidiary = companyName['Subsidiary'].tolist()
        for ind, tmp_word in enumerate(list_of_company):
            subsidiary_name = list_of_subsidiary[ind]
            log.info(
                "PROCESSING - {} COMPANY. OUT OF - {}".format(ind+1, len(list_of_company)))
            print("PROCESSING - ", ind+1,
                  "COMPANY.OUT OF - ", len(list_of_company))
            # Below utility will create combination of company name
            search = (ExcelUtils.search_text_combination(tmp_word.split()))
            search_input = search[::-1]
            for i in search_input:
                gtotal = gtotal+1
                log.info(
                    "FOR SEARCH KEY : {} || TOTAL NUMBER OF ITERATION {} ----- ".format(i, gtotal))

                print("FOR SEARCH KEY :", i,
                      "|| TOTAL NUMBER OF ITERATION==>", gtotal)
                temp = i
                search_area = driver.find_element(
                    By.ID, "ctl00_ContentPlaceHolder1_textbox2")
                # Clear the text box
                search_area.clear()
                # Setting the name of company to search area
                search_area.send_keys(temp)
                # clicking the button for search
                gobutton = driver.find_element(
                    By.ID, "ctl00_ContentPlaceHolder1_btn")
                driver.execute_script("arguments[0].click();", gobutton)
                driver.refresh()
                # identify the number of tables
                table_count = len(driver.find_elements(
                    By.XPATH, '//table[@class="table Grid1"]/tbody/tr/td/table'))
                if table_count != 0:
                    # counting the number of columns in the table
                    no_data_cols = len(driver.find_elements(
                        By.XPATH, '//table[@class="table Grid1"]/tbody/tr[3]/td'))
                    # counting the number of pages from pagination
                    page_count = len(driver.find_elements(
                        By.XPATH, '//table[@class="table Grid1"]/tbody/tr[1]/td/table/tbody/tr/td/a'))+1
                else:
                    # counting the number of columns in the table
                    no_data_cols = len(driver.find_elements(
                        By.XPATH, '//table[@class="table Grid1"]/tbody/tr[2]/td'))
                    # setting  default  pages number
                    page_count = 1

                print("page_count=== ", page_count)
                if no_data_cols == 1:
                    # NO data found,skipping this iteration
                    continue
                # Scrapping the data
                a = collectData(driver, table_count, tmp_word,
                                page_count, ws, subsidiary_name)
                # if rflag value is 0 then company name does not matches
                if a == 0:
                    log.info("NO MATCH FOUND FOR ----- {}".format(tmp_word))
                    break

    except Exception as e:
        err_log.error("ERRRO ----- {}".format(str(e)))
        print('THERE IS SOME ERROR--- ' + str(e))


def dumpToDatabase(path, row_count):
    log.info("** -- CONNECTING TO DATABASE -- **")
    conn = pymssql.connect(
        host=r'101.53.148.84',
        port=r'8649',
        user=r'user_at',
        password=r'Ye4@83kv3',
        database='db_at'
    )
    cursor = conn.cursor(as_dict=True)
    log.info("** -- DATABASE CONNECTION ESTABLISHED SUCCESSFULLY -- **")

    # workbook object is created
    wb_obj = load_workbook(path)

    sheet_obj = wb_obj.active
    max_row = len(sheet_obj['A'])
    max_col = row_count
    l = []
    for r in range(2, max_row+1):
        col = []
        for c in range(1, max_col + 1):
            cell_obj = sheet_obj.cell(row=r, column=c)
            col.insert(c, cell_obj.value)
            # print(cell_obj.value, end = ",")
        l.append(tuple(col))
    # print(l)

    log.info("INSERTION OF {} Records is STARTED Please Wait !!!". format(max_row))
    start_time = time.time()
    q = "insert into parivesh_central(Proposal_No, MOEFCC_File_No, Project_Name, Company, Proposal_Status, Location,Important_Dates, Category,Company_or_Proponent,Type_of_project, Attached_Files,Acknowlegment,Pagination,input_company_name,subsidiary_name) values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
    cursor.executemany(q, l)
    conn.commit()
    log.info("DATA IS COMMITTED TO DATABASE. Time Taken to Store {} Record is {}".format(
        max_row, time.time()-start_time))
    conn.close()


# code start from here
start_time = time.time()
x = datetime.datetime.now()
# Here we are defining the name of the file...
today = "centeralData_"+str(x.day) + str(x.month) + str(x.year)+"_" + \
    str(x.strftime("%H"))+str(x.strftime("%M"))+str(x.strftime("%S"))
ROOT_DIR = os.path.dirname(os.path.abspath(__file__))
# To get the Operating System
plat_os = platform.uname()[0]
folder_path = ''
log_path = ''
error_path = ''
# Below logic will create folder in project root directory based on the operating system where we are running the code
if plat_os == 'Windows':
    folder_path = ROOT_DIR+"\\central_output\\"
    log_path = ROOT_DIR+"\\central_log\\"
    error_path = ROOT_DIR+"\\error_log\\"
else:
    folder_path = ROOT_DIR+"/central_output/"
    log_path = ROOT_DIR+"/central_log/"
    error_path = ROOT_DIR+"/error_log/"


if not os.path.exists(folder_path):
    os.makedirs(folder_path)
if not os.path.exists(log_path):
    os.makedirs(log_path)
if not os.path.exists(error_path):
    os.makedirs(error_path)

print(log_path+today+'.log')
# LOG_FORMAT = '{lineno} ***  {asctime} ----- {message}'
# log.basicConfig(filename=log_path+today+'.log',
#                 style='{', format=LOG_FORMAT, level=log.DEBUG)

# Creating INFO LOG file
formatter = logging.Formatter('%(asctime)s %(levelname)s %(message)s')


def setup_logger(name, log_file, level=logging.INFO):
    handler = logging.FileHandler(log_file)
    handler.setFormatter(formatter)
    logger = logging.getLogger(name)
    logger.setLevel(level)
    logger.addHandler(handler)
    return logger


# Info  File logger
log = setup_logger('first_logger', log_path+today+'.log')
log.info('This is just info Log')

# Error file logger
err_log = setup_logger('second_logger', error_path+today+'.log')
err_log.error('This is an error Log')


path = folder_path+today+".xlsx"
workbook = Workbook()
workbook.save(path)
wb = load_workbook(filename=path)
ws = wb['Sheet']
# Here we are defining the header name of excel file and saving it.
header_list = ["Proposal No", "MOEFCC File No", "Project Name", "Company", "Proposal Status", "Location",
               "Important Dates", "Category", "Company/Proponent", "Type of project", "Attached Files", "Acknowlegment", "Pagination", "Input Company Name", "Subsidiary Name"]
ws.append(header_list)
wb.save(path)
# This function will start scrapping the website data and store it in excel file
centralDataScrapping(ws)
# this function will dump the excel data into database
dumpToDatabase(path, len(header_list))
log.info("TOTAL TIME ----- {}".format(time.time()-start_time))
print("TOTAL TIME========>", time.time()-start_time)
